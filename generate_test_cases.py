"""
Generate comprehensive test case Excel sheets for GiveAwayLead.
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Color palette ──────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":   "4F2D8A",   # dark purple
    "header_font": "FFFFFF",
    "high":        "FF4D4D",
    "medium":      "FFA500",
    "low":         "4CAF50",
    "positive":    "D6F5D6",
    "negative":    "FFE0E0",
    "edge":        "FFF3CC",
    "section_bg":  "EDE7F6",
    "alt_row":     "F9F6FF",
    "white":       "FFFFFF",
    "pass":        "C8E6C9",
    "fail":        "FFCDD2",
    "blocked":     "FFE0B2",
}

HEADERS = [
    "TC ID", "Module", "Sub-Module", "Test Case Name",
    "Description / Test Steps", "Pre-Conditions",
    "Expected Result", "Actual Result", "Status",
    "Priority", "Type", "Notes",
]

COL_WIDTHS = [10, 18, 20, 34, 62, 34, 44, 24, 10, 10, 12, 22]

PRIORITY = {"H": "High", "M": "Medium", "L": "Low"}
TYPE     = {"P": "Positive", "N": "Negative", "E": "Edge Case"}


# ── Styling helpers ────────────────────────────────────────────────────────────

def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def thin_border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws):
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill    = hex_fill(COLORS["header_bg"])
        cell.font    = Font(bold=True, color=COLORS["header_font"], size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

def write_row(ws, row_num: int, values: list, tc_type: str = "P"):
    bg = {"P": COLORS["positive"], "N": COLORS["negative"], "E": COLORS["edge"]}
    row_fill = hex_fill(bg.get(tc_type, COLORS["white"]) if row_num % 2 == 0 else COLORS["alt_row"])

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border    = thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Type-based background only on even rows
        if col_idx not in (9, 10, 11):  # status/priority/type get special color
            if row_num % 2 == 0:
                cell.fill = row_fill

        # Priority color
        if col_idx == 10:
            p = val or ""
            if p == "High":   cell.fill = hex_fill(COLORS["high"]);   cell.font = Font(bold=True, color="FFFFFF")
            elif p == "Medium": cell.fill = hex_fill(COLORS["medium"]); cell.font = Font(bold=True, color="FFFFFF")
            elif p == "Low":  cell.fill = hex_fill(COLORS["low"]);    cell.font = Font(bold=True, color="FFFFFF")

        # Type color chip
        if col_idx == 11:
            t = val or ""
            if t == "Positive":  cell.fill = hex_fill("D6F5D6"); cell.font = Font(bold=True, color="1B5E20")
            elif t == "Negative": cell.fill = hex_fill("FFE0E0"); cell.font = Font(bold=True, color="B71C1C")
            elif t == "Edge Case": cell.fill = hex_fill("FFF3CC"); cell.font = Font(bold=True, color="795548")

    ws.row_dimensions[row_num].height = 60


def tc(tc_id, module, sub, name, steps, precond, expected, priority="M", ttype="P", notes=""):
    return [
        tc_id, module, sub, name, steps, precond, expected,
        "", "", PRIORITY[priority], TYPE[ttype], notes,
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  BACKEND TEST CASES
# ══════════════════════════════════════════════════════════════════════════════

BACKEND_CASES = [
    # ── AUTH ──────────────────────────────────────────────────────────────────
    tc("BE-AUTH-001","Auth","Register","Register with valid data",
       "POST /api/auth/register\nBody: {name, email, phone, password, role:'user'}",
       "Backend running, MongoDB connected",
       "HTTP 201\nUser created in DB\nJWT token returned\nrole='user'","H","P"),

    tc("BE-AUTH-002","Auth","Register","Register as seller",
       "POST /api/auth/register\nBody: {name, email, phone, password, role:'seller'}",
       "Backend running",
       "HTTP 201\nrole='seller'\nreferral_code generated","H","P"),

    tc("BE-AUTH-003","Auth","Register","Duplicate email registration",
       "POST /api/auth/register twice with same email",
       "User already exists with that email",
       "HTTP 409 Conflict\n'Email already registered' error","H","N"),

    tc("BE-AUTH-004","Auth","Register","Missing required fields",
       "POST /api/auth/register\nOmit 'email' from body",
       "Backend running",
       "HTTP 422 Unprocessable Entity\nValidation error details returned","M","N"),

    tc("BE-AUTH-005","Auth","Register","Weak/empty password",
       "POST /api/auth/register\npassword: ''",
       "Backend running",
       "HTTP 422 or 400\nPassword validation error","M","E"),

    tc("BE-AUTH-006","Auth","Login","Login with correct credentials",
       "POST /api/auth/login\nBody: {email, password}",
       "User exists in DB",
       "HTTP 200\ntoken in response\nSet-Cookie: token=...; HttpOnly","H","P"),

    tc("BE-AUTH-007","Auth","Login","Login with wrong password",
       "POST /api/auth/login\nBody: {email, wrong_password}",
       "User exists in DB",
       "HTTP 401\n'Invalid credentials' error","H","N"),

    tc("BE-AUTH-008","Auth","Login","Login with non-existent email",
       "POST /api/auth/login\nBody: {unknown@test.com, password}",
       "Email not in DB",
       "HTTP 401\n'Invalid credentials' (not 'user not found' — no enumeration)","H","N"),

    tc("BE-AUTH-009","Auth","Me","Get current user profile",
       "GET /api/auth/me\nHeader: Authorization: Bearer <valid_token>",
       "User logged in, valid JWT",
       "HTTP 200\nReturns id, name, email, role, referral_code","M","P"),

    tc("BE-AUTH-010","Auth","Me","Access /me without token",
       "GET /api/auth/me\nNo Authorization header",
       "No token",
       "HTTP 401 or 403","H","N"),

    tc("BE-AUTH-011","Auth","Me","Access /me with expired token",
       "GET /api/auth/me\nHeader: Authorization: Bearer <expired_jwt>",
       "Token expired",
       "HTTP 401\n'Invalid token' error","H","N"),

    tc("BE-AUTH-012","Auth","Profile","Update profile successfully",
       "PUT /api/auth/profile\nBody: {name:'New Name', phone:'9999999999', city:'Delhi'}",
       "Valid JWT in header",
       "HTTP 200\nDB updated\nNew values returned","M","P"),

    tc("BE-AUTH-013","Auth","Logout","Logout clears cookie",
       "POST /api/auth/logout",
       "User logged in",
       "HTTP 200\nSet-Cookie: token=; Max-Age=0 (cookie cleared)","M","P"),

    tc("BE-AUTH-014","Auth","Role","User cannot access seller endpoint",
       "GET /api/campaigns/mine\nJWT with role='user'",
       "User JWT",
       "HTTP 403 Forbidden","H","N"),

    tc("BE-AUTH-015","Auth","Role","User cannot access admin endpoint",
       "GET /api/admin/users\nJWT with role='user'",
       "User JWT",
       "HTTP 403 Forbidden","H","N"),

    # ── CAMPAIGNS ─────────────────────────────────────────────────────────────
    tc("BE-CAMP-001","Campaigns","List","List active campaigns",
       "GET /api/campaigns",
       "Active campaigns in DB",
       "HTTP 200\nArray of CampaignOut objects\nOnly status='active' returned","H","P"),

    tc("BE-CAMP-002","Campaigns","List","Filter by category",
       "GET /api/campaigns?category=Electronics",
       "Active campaigns exist in Electronics category",
       "HTTP 200\nAll returned campaigns have category='Electronics'","M","P"),

    tc("BE-CAMP-003","Campaigns","List","Filter by invalid category",
       "GET /api/campaigns?category=DoesNotExist",
       "Backend running",
       "HTTP 200\nEmpty array []","M","E"),

    tc("BE-CAMP-004","Campaigns","List","Featured campaigns limited to 3",
       "GET /api/campaigns?featured=true",
       "More than 3 active campaigns in DB",
       "HTTP 200\nArray length ≤ 3","M","P"),

    tc("BE-CAMP-005","Campaigns","Detail","Get campaign by valid ID",
       "GET /api/campaigns/{valid_id}",
       "Campaign exists",
       "HTTP 200\nAll CampaignDetail fields returned\nparticipants count is real-time","H","P"),

    tc("BE-CAMP-006","Campaigns","Detail","Get campaign with invalid ObjectID",
       "GET /api/campaigns/not_an_objectid",
       "Backend running",
       "HTTP 400 'Invalid campaign ID'","M","N"),

    tc("BE-CAMP-007","Campaigns","Detail","Get non-existent campaign",
       "GET /api/campaigns/507f1f77bcf86cd799439011",
       "ID is valid format but not in DB",
       "HTTP 404 'Campaign not found'","M","N"),

    tc("BE-CAMP-008","Campaigns","Create","Seller creates campaign",
       "POST /api/campaigns\nBody: {title, price, description, category, winners:1, duration_days:7, offer_type:'free'}\nSeller JWT",
       "Valid seller JWT",
       "HTTP 201\nCampaign created with status='pending'\nseller_id set correctly","H","P"),

    tc("BE-CAMP-009","Campaigns","Create","User cannot create campaign",
       "POST /api/campaigns\nUser JWT",
       "User JWT (role='user')",
       "HTTP 403 Forbidden","H","N"),

    tc("BE-CAMP-010","Campaigns","Create","Missing required fields",
       "POST /api/campaigns\nBody omits 'title'\nSeller JWT",
       "Valid seller JWT",
       "HTTP 422 Unprocessable Entity","M","N"),

    tc("BE-CAMP-011","Campaigns","Approve","Admin approves pending campaign",
       "PATCH /api/campaigns/{id}/approve\nAdmin JWT",
       "Campaign with status='pending'",
       "HTTP 200 {status:'approved'}\nCampaign status becomes 'active'\nSeller notified by email (if RESEND configured)","H","P"),

    tc("BE-CAMP-012","Campaigns","Approve","Non-admin cannot approve",
       "PATCH /api/campaigns/{id}/approve\nSeller JWT",
       "Campaign in pending",
       "HTTP 403 Forbidden","H","N"),

    tc("BE-CAMP-013","Campaigns","Reject","Admin rejects pending campaign",
       "PATCH /api/campaigns/{id}/reject\nAdmin JWT",
       "Campaign with status='pending'",
       "HTTP 200 {status:'rejected'}\nCampaign status='rejected'","H","P"),

    tc("BE-CAMP-014","Campaigns","Draw","Draw winners on active campaign",
       "POST /api/campaigns/{id}/draw\nAdmin JWT",
       "Campaign active with entries",
       "HTTP 200\nArray of WinnerOut objects\nWinners' entry_status='Won'\nOthers entry_status='Lost'\nCampaign status='ended', winner_drawn=True","H","P"),

    tc("BE-CAMP-015","Campaigns","Draw","Draw on campaign with no entries",
       "POST /api/campaigns/{id}/draw\nAdmin JWT",
       "Campaign has 0 entries",
       "HTTP 200\nEmpty array []","M","E"),

    tc("BE-CAMP-016","Campaigns","Draw","Draw twice on same campaign",
       "POST /api/campaigns/{id}/draw twice\nAdmin JWT",
       "Campaign already drawn",
       "HTTP 409 'Winners already drawn'","H","N"),

    tc("BE-CAMP-017","Campaigns","Mine","Seller gets own campaigns",
       "GET /api/campaigns/mine\nSeller JWT",
       "Seller has campaigns",
       "HTTP 200\nOnly campaigns belonging to this seller_id","M","P"),

    tc("BE-CAMP-018","Campaigns","Pending","Admin gets pending campaigns",
       "GET /api/campaigns/pending\nAdmin JWT",
       "Pending campaigns exist",
       "HTTP 200\nAll status='pending'","M","P"),

    # ── ENTRIES ───────────────────────────────────────────────────────────────
    tc("BE-ENT-001","Entries","Create","User enters active campaign",
       "POST /api/entries\nBody: {campaign_id, name, email, phone, city}\nUser JWT",
       "Active campaign exists, user not already entered",
       "HTTP 201\nEntry created\nCampaign participants incremented by 1","H","P"),

    tc("BE-ENT-002","Entries","Create","Duplicate entry same campaign",
       "POST /api/entries twice with same user+campaign",
       "User already entered this campaign",
       "HTTP 409 'Already entered this campaign'","H","N"),

    tc("BE-ENT-003","Entries","Create","Enter inactive (pending) campaign",
       "POST /api/entries\ncampaign_id of a 'pending' campaign",
       "Campaign not active",
       "HTTP 400 'Campaign is not accepting entries'","H","N"),

    tc("BE-ENT-004","Entries","Create","Enter with valid referral code",
       "POST /api/entries\nBody includes ref_code of another user",
       "Referrer user with that referral_code exists",
       "HTTP 201\nEntry has referred_by=ref_code\nreferrer_name populated","M","P"),

    tc("BE-ENT-005","Entries","Create","Self-referral rejected",
       "POST /api/entries\nref_code = own user's referral_code",
       "User uses own ref code",
       "HTTP 201 but referred_by='' (self-referral silently ignored)","M","E"),

    tc("BE-ENT-006","Entries","Create","Invalid campaign ID",
       "POST /api/entries\ncampaign_id='notanid'",
       "Backend running",
       "HTTP 400 'Invalid campaign ID'","M","N"),

    tc("BE-ENT-007","Entries","My Entries","Get current user's entries",
       "GET /api/entries/me\nUser JWT",
       "User has entries",
       "HTTP 200\nArray of EntryOut for this user only","M","P"),

    tc("BE-ENT-008","Entries","Winners","Public winners list",
       "GET /api/entries/winners",
       "Some entries with status='Won' exist",
       "HTTP 200\nArray with name, campaign_title, prize, city\nNo auth required","H","P"),

    tc("BE-ENT-009","Entries","Winners","Winners list is empty",
       "GET /api/entries/winners",
       "No entries with status='Won'",
       "HTTP 200\nEmpty array []","M","E"),

    tc("BE-ENT-010","Entries","Leaderboard","Referral leaderboard for campaign",
       "GET /api/entries/campaign/{id}/leaderboard",
       "Campaign with referral entries",
       "HTTP 200\nSorted by count desc\nMax 10 results","M","P"),

    tc("BE-ENT-011","Entries","CSV","Export leads as CSV (seller)",
       "GET /api/entries/campaign/{id}/csv\nSeller JWT (own campaign)",
       "Campaign has entries, seller owns it",
       "HTTP 200\nContent-Type: text/csv\nCSV has correct columns: Name,Email,Phone,City,Campaign,Joined At,Status,Referred By","H","P"),

    tc("BE-ENT-012","Entries","CSV","Seller cannot export another seller's leads",
       "GET /api/entries/campaign/{other_seller_campaign_id}/csv\nSeller JWT",
       "Campaign belongs to different seller",
       "HTTP 403 Forbidden","H","N"),

    # ── ADMIN ─────────────────────────────────────────────────────────────────
    tc("BE-ADM-001","Admin","Users","List all users",
       "GET /api/admin/users\nAdmin JWT",
       "Users in DB",
       "HTTP 200\n{total, users: [...]}\npassword_hash excluded","H","P"),

    tc("BE-ADM-002","Admin","Users","Search users by name",
       "GET /api/admin/users?search=John\nAdmin JWT",
       "Users with 'John' in name exist",
       "HTTP 200\nFiltered results","M","P"),

    tc("BE-ADM-003","Admin","Users","Ban user",
       "PATCH /api/admin/users/{id}/ban\nAdmin JWT",
       "User exists, not banned",
       "HTTP 200 {status:'banned'}\nis_banned=True in DB","H","P"),

    tc("BE-ADM-004","Admin","Users","Unban user",
       "PATCH /api/admin/users/{id}/unban\nAdmin JWT",
       "User is banned",
       "HTTP 200 {status:'unbanned'}\nis_banned=False in DB","H","P"),

    tc("BE-ADM-005","Admin","Analytics","Platform analytics",
       "GET /api/admin/analytics\nAdmin JWT",
       "Some entries and users exist",
       "HTTP 200\n{series:[{date,entries,users}] 30 items, top_campaigns:[...]}","M","P"),

    tc("BE-ADM-006","Admin","Analytics","Analytics with no data",
       "GET /api/admin/analytics\nAdmin JWT",
       "Empty DB",
       "HTTP 200\nAll series counts=0, top_campaigns=[]","M","E"),

    tc("BE-ADM-007","Admin","Campaigns","List all campaigns",
       "GET /api/admin/campaigns\nAdmin JWT",
       "Campaigns in DB",
       "HTTP 200\n{total, campaigns:[...]}","M","P"),

    tc("BE-ADM-008","Admin","Campaigns","Filter by status",
       "GET /api/admin/campaigns?status=pending\nAdmin JWT",
       "Pending campaigns exist",
       "HTTP 200\nAll returned have status='pending'","M","P"),

    tc("BE-ADM-009","Admin","Draw","Admin draw endpoint",
       "POST /api/admin/draw/{campaign_id}\nAdmin JWT",
       "Active campaign with entries, not yet drawn",
       "HTTP 200\n{campaign_id, winners:[{entry_id,name,email,city,won_at}]}","H","P"),

    # ── FRAUD ─────────────────────────────────────────────────────────────────
    tc("BE-FRD-001","Fraud","Detection","Detect duplicate phone numbers",
       "GET /api/admin/fraud/suspicious\nAdmin JWT",
       "Two entries with same phone, different user_ids",
       "HTTP 200\nduplicate_phones array contains the group\ncount≥2","H","P"),

    tc("BE-FRD-002","Fraud","Detection","Detect duplicate emails",
       "GET /api/admin/fraud/suspicious\nAdmin JWT",
       "Two entries with same email, different user_ids",
       "HTTP 200\nduplicate_emails array contains the group","H","P"),

    tc("BE-FRD-003","Fraud","Detection","No fraud returns empty",
       "GET /api/admin/fraud/suspicious\nAdmin JWT",
       "All entries have unique phone+email per user",
       "HTTP 200\ntotal_suspicious=0\nBoth arrays empty","M","E"),

    tc("BE-FRD-004","Fraud","Disqualify","Disqualify an entry",
       "PATCH /api/admin/fraud/disqualify/{entry_id}\nAdmin JWT",
       "Entry exists, status='Active'",
       "HTTP 200\nentry_status='Disqualified' in DB","H","P"),

    tc("BE-FRD-005","Fraud","Disqualify","Disqualify non-existent entry",
       "PATCH /api/admin/fraud/disqualify/507f1f77bcf86cd799439011\nAdmin JWT",
       "Entry ID not in DB",
       "HTTP 404 'Entry not found'","M","N"),

    tc("BE-FRD-006","Fraud","Disqualify","Non-admin cannot disqualify",
       "PATCH /api/admin/fraud/disqualify/{entry_id}\nSeller JWT",
       "Valid seller JWT",
       "HTTP 403 Forbidden","H","N"),

    # ── PAYMENTS ──────────────────────────────────────────────────────────────
    tc("BE-PAY-001","Payments","Plans","Get payment plans",
       "GET /api/payments/plans",
       "Backend running (no auth required)",
       "HTTP 200\nArray with Basic (₹999) and Pro (₹2499) plans","H","P"),

    tc("BE-PAY-002","Payments","Order","Create Razorpay order",
       "POST /api/payments/order\nBody: {plan:'basic', campaign_id}\nSeller JWT",
       "RAZORPAY_KEY_ID + SECRET configured in .env",
       "HTTP 200\n{order_id, amount:99900, currency:'INR', key_id, plan_name}","H","P"),

    tc("BE-PAY-003","Payments","Order","Order without Razorpay config",
       "POST /api/payments/order\nBody: {plan:'basic'}\nSeller JWT",
       "RAZORPAY_KEY_ID not set",
       "HTTP 503 'Payment gateway not configured'","H","N"),

    tc("BE-PAY-004","Payments","Order","Order with invalid plan",
       "POST /api/payments/order\nBody: {plan:'enterprise_xyz'}",
       "Seller JWT",
       "HTTP 400 'Invalid plan'","M","N"),

    tc("BE-PAY-005","Payments","Verify","Verify valid payment signature",
       "POST /api/payments/verify\nBody: {razorpay_order_id, razorpay_payment_id, razorpay_signature, campaign_id, plan}\nSeller JWT",
       "HMAC-SHA256 signature matches",
       "HTTP 200 {status:'verified', payment_id}\nPayment inserted in DB\nCampaign moved from 'payment_pending' to 'pending'","H","P"),

    tc("BE-PAY-006","Payments","Verify","Tampered payment signature",
       "POST /api/payments/verify\nrazorpay_signature='tampered_value'",
       "Signature does not match computed HMAC",
       "HTTP 400 'Payment verification failed'","H","N"),

    tc("BE-PAY-007","Payments","History","Seller payment history",
       "GET /api/payments/history\nSeller JWT",
       "Seller has previous payments",
       "HTTP 200\nArray sorted newest-first\nOnly this seller's payments","M","P"),

    tc("BE-PAY-008","Payments","Revenue","Admin revenue dashboard",
       "GET /api/payments/admin/revenue\nAdmin JWT",
       "Payments exist in DB",
       "HTTP 200\n{total_revenue, total_payments, by_plan, recent_payments}","H","P"),

    tc("BE-PAY-009","Payments","Revenue","Non-admin cannot see revenue",
       "GET /api/payments/admin/revenue\nSeller JWT",
       "Seller JWT",
       "HTTP 403 Forbidden","H","N"),

    # ── SCHEDULER ─────────────────────────────────────────────────────────────
    tc("BE-SCH-001","Scheduler","Email","Ending-soon email sent once",
       "Campaign ends_at within 24h\nScheduler runs check_ending_soon()",
       "APScheduler running, campaign active, reminder_sent=False",
       "reminder_sent=True set in DB before email\nEmail fired to all active entrants\nNo duplicate email on second scheduler run","H","P"),

    tc("BE-SCH-002","Scheduler","Email","Reminder not sent twice",
       "Campaign reminder_sent=True in DB\nScheduler runs again",
       "reminder_sent already True",
       "No email sent\nDB unchanged","H","N"),

    tc("BE-SCH-003","Scheduler","Email","No email if RESEND not configured",
       "RESEND_API_KEY not in .env\nScheduler runs",
       "RESEND_API_KEY missing",
       "No error thrown\nScheduler continues silently (graceful skip)","M","E"),

    # ── STATS ─────────────────────────────────────────────────────────────────
    tc("BE-STA-001","Stats","Public","Public platform stats",
       "GET /api/stats",
       "Backend running",
       "HTTP 200\n{total_campaigns, total_entries, total_winners, active_now}\nNo auth required","M","P"),
]


# ══════════════════════════════════════════════════════════════════════════════
#  FRONTEND TEST CASES
# ══════════════════════════════════════════════════════════════════════════════

FRONTEND_CASES = [
    # ── NAVIGATION / LAYOUT ───────────────────────────────────────────────────
    tc("FE-NAV-001","Navigation","Navbar","Navbar renders on public pages",
       "1. Open http://localhost:3000\n2. Inspect navbar",
       "App running",
       "Logo, Campaigns link, Winners link, How it Works link visible\nSign In + Register buttons shown for guests","H","P"),

    tc("FE-NAV-002","Navigation","Navbar","Navbar shows user menu when logged in",
       "1. Login as any role\n2. Inspect navbar",
       "User logged in",
       "User avatar/name shown\nDashboard link visible\nLogout option in dropdown","H","P"),

    tc("FE-NAV-003","Navigation","Mobile","Mobile hamburger menu",
       "1. Resize browser to < 768px\n2. Click hamburger icon",
       "Mobile viewport",
       "Slide-in menu opens\nAll nav links visible\nClicking link closes menu","H","P"),

    tc("FE-NAV-004","Navigation","Sidebar","Sidebar shows correct links per role",
       "1. Login as user → check sidebar\n2. Login as seller → check sidebar\n3. Login as admin → check sidebar",
       "Users of each role exist",
       "User: Dashboard, My Campaigns, My Wins, Profile\nSeller: Dashboard, Create, My Campaigns, Leads, Analytics, Payments\nAdmin: Dashboard, Approvals, Campaigns, Users, Winner Draw, Analytics, Fraud, Revenue","H","P"),

    tc("FE-NAV-005","Navigation","Sidebar","Active link highlighted",
       "1. Navigate to /dashboard/user/campaigns",
       "User logged in",
       "My Campaigns link is highlighted/active in sidebar","M","P"),

    tc("FE-NAV-006","Navigation","Footer","Footer renders",
       "1. Open any public page\n2. Scroll to bottom",
       "App running",
       "Footer visible with links and copyright","L","P"),

    # ── AUTH PAGES ────────────────────────────────────────────────────────────
    tc("FE-AUTH-001","Auth","Login","Login page renders",
       "1. Navigate to /login",
       "App running",
       "Email + Password fields rendered\nSign In button visible\nRegister link visible","H","P"),

    tc("FE-AUTH-002","Auth","Login","Successful login redirects correctly",
       "1. Navigate to /login\n2. Enter valid user credentials\n3. Submit",
       "Valid user account exists",
       "Redirected to /dashboard/user\nSidebar shows user role links","H","P"),

    tc("FE-AUTH-003","Auth","Login","Seller login redirects to seller dashboard",
       "1. Login with seller credentials",
       "Valid seller account",
       "Redirected to /dashboard/seller","H","P"),

    tc("FE-AUTH-004","Auth","Login","Admin login redirects to admin dashboard",
       "1. Login with admin credentials",
       "Valid admin account",
       "Redirected to /dashboard/admin","H","P"),

    tc("FE-AUTH-005","Auth","Login","Wrong password shows error",
       "1. Navigate to /login\n2. Enter correct email, wrong password\n3. Submit",
       "User exists",
       "Error message displayed\nUser stays on /login\nNo redirect","H","N"),

    tc("FE-AUTH-006","Auth","Login","Empty form submission",
       "1. Navigate to /login\n2. Click Sign In without filling fields",
       "Login page open",
       "Browser HTML5 validation or inline error shown\nNo API call made","M","N"),

    tc("FE-AUTH-007","Auth","Register","Register page renders",
       "1. Navigate to /register",
       "App running",
       "Name, Email, Phone, Password, Role selector visible\nSubmit button present","H","P"),

    tc("FE-AUTH-008","Auth","Register","Successful user registration",
       "1. Fill all fields\n2. Select role: User\n3. Submit",
       "Email not already used",
       "Account created\nRedirected to /dashboard/user","H","P"),

    tc("FE-AUTH-009","Auth","Register","Duplicate email shows error",
       "1. Register with already-used email",
       "Email already exists in DB",
       "Error message: 'Email already registered'\nUser stays on /register","H","N"),

    tc("FE-AUTH-010","Auth","Logout","Logout clears session",
       "1. Login\n2. Click Logout in sidebar",
       "User logged in",
       "Redirected to /login or /\nSidebar gone\nCookie cleared","H","P"),

    tc("FE-AUTH-011","Auth","Protection","Dashboard redirects to login if not authenticated",
       "1. Clear cookies\n2. Navigate to /dashboard/user",
       "Not logged in",
       "Redirected to /login","H","N"),

    tc("FE-AUTH-012","Auth","Protection","User cannot access /dashboard/admin",
       "1. Login as user\n2. Navigate to /dashboard/admin",
       "User JWT (role=user)",
       "Redirected or shown 403 page\nNot allowed in admin area","H","N"),

    # ── PUBLIC PAGES ──────────────────────────────────────────────────────────
    tc("FE-PUB-001","Public","Home","Landing page renders",
       "1. Open http://localhost:3000",
       "App running, backend running",
       "Hero section visible\nFeatured campaigns load (real data from API)\nHow it Works section visible\nTestimonials visible","H","P"),

    tc("FE-PUB-002","Public","Home","Featured campaigns load from API",
       "1. Open home page\n2. Check featured campaigns grid",
       "Active campaigns in DB",
       "Up to 3 real campaign cards shown\nEach has image, title, brand, participant count","H","P"),

    tc("FE-PUB-003","Public","Campaigns","Campaigns listing page",
       "1. Navigate to /campaigns",
       "Active campaigns in DB",
       "Campaign grid loaded\nEach card shows image, title, price, time left, participants","H","P"),

    tc("FE-PUB-004","Public","Campaigns","Category filter works",
       "1. Navigate to /campaigns\n2. Click 'Electronics' category filter",
       "Active campaigns in multiple categories",
       "Only Electronics campaigns shown\nOther categories hidden","H","P"),

    tc("FE-PUB-005","Public","Campaigns","Search campaigns",
       "1. Navigate to /campaigns\n2. Type in search box",
       "Campaigns with different titles",
       "Results filter in real-time or on submit\nOnly matching campaigns shown","M","P"),

    tc("FE-PUB-006","Public","Campaign Detail","Campaign detail page renders",
       "1. Click any campaign card",
       "Active campaign exists",
       "Image, title, brand, stats (entries, winners, time left) visible\nTerms & conditions shown\nEntry form rendered","H","P"),

    tc("FE-PUB-007","Public","Campaign Detail","OG meta tags present",
       "1. View page source of /campaigns/{id}",
       "Campaign exists",
       "<meta property='og:title'> with campaign title\n<meta property='og:image'> with campaign image\n<meta name='twitter:card'> present","M","P"),

    tc("FE-PUB-008","Public","Campaign Detail","Top referrers leaderboard shown",
       "1. Open campaign with referral entries",
       "Referral entries exist",
       "Top Referrers section visible\nNames and referral counts shown\nMedal colors for top 3","M","P"),

    tc("FE-PUB-009","Public","Winners","Winners page renders",
       "1. Navigate to /winners",
       "Some entries with status='Won'",
       "Winner cards shown\nNames slightly masked (first name + last initial)\nCity shown\nConfetti fires on load","H","P"),

    tc("FE-PUB-010","Public","Winners","Empty winners page",
       "1. Navigate to /winners with no winners in DB",
       "No Won entries",
       "Empty state message shown\n'Browse Giveaways' CTA visible\nNo confetti","M","E"),

    tc("FE-PUB-011","Public","Embed","Embed widget renders",
       "1. Navigate to /embed/{valid_campaign_id}",
       "Active campaign exists",
       "Campaign card shown\nImage, title, stats (entries, winners, time left)\n'Enter to Win' CTA links to full campaign","H","P"),

    tc("FE-PUB-012","Public","Embed","Embed fits in iframe",
       "1. Create HTML page with <iframe src='/embed/{id}' width='360' height='400'>\n2. Open in browser",
       "Embed page accessible",
       "Widget renders inside iframe without overflow or scrollbars","M","P"),

    # ── ENTRY FORM ────────────────────────────────────────────────────────────
    tc("FE-ENT-001","Entry Form","Render","Entry form shown to logged-in users",
       "1. Login as user\n2. Open any active campaign",
       "User logged in, active campaign",
       "Entry form visible with Name, Email, Phone, City fields\nSubmit button active","H","P"),

    tc("FE-ENT-002","Entry Form","Render","Entry form shown to guests (pre-filled login prompt)",
       "1. Open campaign without logging in",
       "Not logged in",
       "Login/Register prompt shown or form rendered\n(Behavior depends on implementation)","H","P"),

    tc("FE-ENT-003","Entry Form","Submit","Successful entry submission",
       "1. Login as user\n2. Open active campaign\n3. Fill form\n4. Submit",
       "User not already entered",
       "Success state shown\nShare panel with referral link visible\nCopy link, WhatsApp, Twitter buttons visible\nParticipant count increments","H","P"),

    tc("FE-ENT-004","Entry Form","Submit","Duplicate entry shows error",
       "1. Enter same campaign twice with same user",
       "Already entered",
       "409 error handled gracefully\n'Already entered' message shown\nShare panel shown (user can still share)","H","N"),

    tc("FE-ENT-005","Entry Form","Referral","Referral code tracked from URL",
       "1. Open /campaigns/{id}?ref=USERCODE\n2. Submit entry form",
       "Referrer has that referral_code",
       "Entry created with referred_by=USERCODE in DB\nReferrer name on leaderboard increments","H","P"),

    tc("FE-ENT-006","Entry Form","Validation","Empty form submission",
       "1. Open campaign\n2. Click Submit without filling fields",
       "Entry form visible",
       "Validation errors shown for required fields\nNo API call made","M","N"),

    tc("FE-ENT-007","Entry Form","Share","Copy referral link",
       "1. Submit entry successfully\n2. Click 'Copy Link' in share panel",
       "Entry submitted successfully",
       "Link copied to clipboard\nButton shows 'Copied!' feedback","M","P"),

    tc("FE-ENT-008","Entry Form","Share","WhatsApp share button",
       "1. Submit entry\n2. Click WhatsApp share",
       "Entry submitted",
       "WhatsApp opens (or new tab) with pre-filled message containing referral link","M","P"),

    # ── USER DASHBOARD ────────────────────────────────────────────────────────
    tc("FE-USR-001","User Dashboard","Overview","User dashboard stats",
       "1. Login as user\n2. Open /dashboard/user",
       "User has entries",
       "Stats cards show: campaigns entered, total wins\nRecent activity visible","H","P"),

    tc("FE-USR-002","User Dashboard","Campaigns","My Campaigns page",
       "1. Navigate to /dashboard/user/campaigns",
       "User has entered campaigns",
       "Active entries shown as cards\nPast/ended entries shown in table\nDraw date and status visible","H","P"),

    tc("FE-USR-003","User Dashboard","Campaigns","Empty state - no campaigns",
       "1. Login as new user with no entries\n2. Open /dashboard/user/campaigns",
       "User has no entries",
       "Empty state message and 'Browse Campaigns' CTA shown","M","E"),

    tc("FE-USR-004","User Dashboard","Wins","My Wins page",
       "1. Navigate to /dashboard/user/wins",
       "User has Won entries",
       "Won entries shown as cards\nPrize and campaign title displayed","H","P"),

    tc("FE-USR-005","User Dashboard","Wins","No wins yet state",
       "1. Login as user with no wins\n2. Open /dashboard/user/wins",
       "User has no Won entries",
       "Empty state with encouraging message and CTA","M","E"),

    tc("FE-USR-006","User Dashboard","Profile","Profile page shows current data",
       "1. Navigate to /dashboard/user/profile",
       "User logged in",
       "Current name, phone, city pre-filled in form","H","P"),

    tc("FE-USR-007","User Dashboard","Profile","Profile update saves changes",
       "1. Edit name on profile page\n2. Click Save",
       "User logged in",
       "Success toast/message shown\nReloading page shows updated name","H","P"),

    # ── SELLER DASHBOARD ──────────────────────────────────────────────────────
    tc("FE-SEL-001","Seller Dashboard","Overview","Seller dashboard stats",
       "1. Login as seller\n2. Open /dashboard/seller",
       "Seller has campaigns",
       "Stats: total campaigns, total leads, active campaigns\nTop campaigns list visible","H","P"),

    tc("FE-SEL-002","Seller Dashboard","Create Campaign","Campaign creation form",
       "1. Navigate to /dashboard/seller/campaigns/new",
       "Logged in as seller",
       "Title, price, description, category, winners, duration, offer type fields visible\nSubmit button renders","H","P"),

    tc("FE-SEL-003","Seller Dashboard","Create Campaign","Create campaign successfully",
       "1. Fill campaign creation form\n2. Submit",
       "Valid seller session",
       "Campaign created with status='pending'\nRedirected to My Campaigns\nNew campaign appears with 'Pending' badge","H","P"),

    tc("FE-SEL-004","Seller Dashboard","Create Campaign","Form validation on submit",
       "1. Submit empty campaign form",
       "Seller dashboard",
       "Required field errors shown\nNo API call made","M","N"),

    tc("FE-SEL-005","Seller Dashboard","My Campaigns","Campaigns list with status badges",
       "1. Navigate to /dashboard/seller/campaigns",
       "Seller has campaigns in various statuses",
       "All campaigns listed\nCorrect badge colors: pending=amber, active=green, ended=grey\nLead count shown per campaign","H","P"),

    tc("FE-SEL-006","Seller Dashboard","Leads","Leads table",
       "1. Navigate to /dashboard/seller/leads\n2. Click on a campaign",
       "Campaign has entries",
       "Table shows Name, Email, Phone, City, Joined At, Status columns\nPagination if >50 rows","H","P"),

    tc("FE-SEL-007","Seller Dashboard","Leads","CSV export download",
       "1. Navigate to leads page\n2. Click Export CSV",
       "Campaign has entries",
       "CSV file downloads\nFile contains correct headers and data","H","P"),

    tc("FE-SEL-008","Seller Dashboard","Analytics","Analytics page renders",
       "1. Navigate to /dashboard/seller/analytics",
       "Seller has campaigns with entries",
       "Stat cards visible\nCSS bar chart per campaign\nNo broken layout","M","P"),

    tc("FE-SEL-009","Seller Dashboard","Payments","Payment plans shown",
       "1. Navigate to /dashboard/seller/payments",
       "Backend running",
       "Basic (₹999) and Pro (₹2499) plan cards visible\nPay buttons enabled","H","P"),

    tc("FE-SEL-010","Seller Dashboard","Payments","Razorpay checkout opens",
       "1. Click Pay ₹999 on Basic plan",
       "RAZORPAY_KEY_ID configured",
       "Razorpay modal/popup opens\nCorrect amount (₹999) shown in checkout","H","P"),

    tc("FE-SEL-011","Seller Dashboard","Payments","Payment history shows after payment",
       "1. Complete a test payment\n2. Check payment history",
       "Payment completed and verified",
       "Payment record appears in history with plan, amount, date","H","P"),

    tc("FE-SEL-012","Seller Dashboard","Payments","No Razorpay config shows error",
       "1. Click Pay without RAZORPAY_KEY_ID set",
       "Razorpay not configured",
       "Error message: 'Payment gateway not configured' or similar\nNo crash","M","E"),

    # ── ADMIN DASHBOARD ───────────────────────────────────────────────────────
    tc("FE-ADM-001","Admin Dashboard","Overview","Admin dashboard stats",
       "1. Login as admin\n2. Open /dashboard/admin",
       "Admin session",
       "Total users, campaigns, entries stats shown\nRecent signups table","H","P"),

    tc("FE-ADM-002","Admin Dashboard","Approvals","Pending campaign approvals",
       "1. Navigate to /dashboard/admin/approvals",
       "Pending campaigns exist",
       "Campaign cards shown\nApprove and Reject buttons visible per campaign","H","P"),

    tc("FE-ADM-003","Admin Dashboard","Approvals","Approve campaign",
       "1. Click Approve on a pending campaign",
       "Pending campaign",
       "Campaign card disappears from approvals list\nCampaign now appears as Active in public listing","H","P"),

    tc("FE-ADM-004","Admin Dashboard","Approvals","Reject campaign",
       "1. Click Reject on a pending campaign",
       "Pending campaign",
       "Campaign removed from approvals queue\nStatus changed to rejected","H","P"),

    tc("FE-ADM-005","Admin Dashboard","Campaigns","Admin campaign list with filters",
       "1. Navigate to /dashboard/admin/campaigns\n2. Click status tabs",
       "Campaigns in various statuses",
       "Tab filters work: All/Active/Pending/Ended/Rejected\nSearch filters by title","H","P"),

    tc("FE-ADM-006","Admin Dashboard","Users","Admin users list",
       "1. Navigate to /dashboard/admin/users",
       "Users exist in DB",
       "User table with Name, Email, Role, Entries, Status\nBan/Unban buttons per user","H","P"),

    tc("FE-ADM-007","Admin Dashboard","Users","Ban user",
       "1. Click Ban on a user",
       "User not currently banned",
       "Button changes to Unban\nBanned badge shown\nUser cannot login","H","P"),

    tc("FE-ADM-008","Admin Dashboard","Users","Search users",
       "1. Type in search box on users page",
       "Multiple users in DB",
       "Table filters to matching users\nOther users hidden","M","P"),

    tc("FE-ADM-009","Admin Dashboard","Winner Draw","Draw page loads campaigns",
       "1. Navigate to /dashboard/admin/draw",
       "Active campaigns with entries",
       "Campaign cards shown with participant count\nDraw Winners button enabled","H","P"),

    tc("FE-ADM-010","Admin Dashboard","Winner Draw","Run draw successfully",
       "1. Click Draw Winners on an active campaign",
       "Campaign has ≥1 active entries",
       "Winners revealed below campaign card\nName + email shown\nCampaign badge changes to 'drawn'\nButton disabled (can't draw twice)","H","P"),

    tc("FE-ADM-011","Admin Dashboard","Winner Draw","Draw with no entries",
       "1. Click Draw Winners on campaign with 0 entries",
       "Campaign has no entries",
       "Empty winners list shown\nNo crash","M","E"),

    tc("FE-ADM-012","Admin Dashboard","Analytics","Analytics charts render",
       "1. Navigate to /dashboard/admin/analytics",
       "Some entries and users in last 30 days",
       "Two line charts visible (entries/day, users/day)\nHorizontal bar chart: top campaigns\nNo broken layout","H","P"),

    tc("FE-ADM-013","Admin Dashboard","Analytics","Analytics with no data",
       "1. Open analytics on empty DB",
       "Empty database",
       "Charts render with all-zero data\nNo JS errors","M","E"),

    tc("FE-ADM-014","Admin Dashboard","Fraud","Fraud page with no suspicious entries",
       "1. Navigate to /dashboard/admin/fraud",
       "All entries unique",
       "Green shield icon shown\n'No suspicious activity detected' message","M","E"),

    tc("FE-ADM-015","Admin Dashboard","Fraud","Fraud page shows suspicious groups",
       "1. Navigate to /dashboard/admin/fraud",
       "Duplicate phone/email entries exist",
       "Warning banner shows count\nGroups listed with shared value + entries\nDisqualify buttons per entry","H","P"),

    tc("FE-ADM-016","Admin Dashboard","Fraud","Disqualify removes entry from UI",
       "1. Click Disqualify on a fraudulent entry",
       "Fraud groups visible",
       "Entry removed from UI immediately\nIf group becomes 0 entries, group removed","H","P"),

    tc("FE-ADM-017","Admin Dashboard","Revenue","Revenue page stats",
       "1. Navigate to /dashboard/admin/revenue",
       "Payments in DB",
       "Total revenue, total payments, avg per payment shown\nBy-plan breakdown shown\nRecent payments table","H","P"),

    tc("FE-ADM-018","Admin Dashboard","Revenue","Revenue page empty state",
       "1. Open revenue page with no payments",
       "No payments in DB",
       "Stats show ₹0\n'No payments yet' message shown","M","E"),

    # ── SEO / PWA ─────────────────────────────────────────────────────────────
    tc("FE-SEO-001","SEO","Sitemap","Sitemap accessible",
       "1. Navigate to http://localhost:3000/sitemap.xml",
       "App running with active campaigns",
       "Valid XML sitemap returned\nContains / /campaigns /winners\nContains /campaigns/{id} for each active campaign","H","P"),

    tc("FE-SEO-002","SEO","Manifest","PWA manifest accessible",
       "1. Navigate to http://localhost:3000/manifest.json",
       "App running",
       "Valid JSON returned\nname, short_name, icons, start_url, theme_color present","M","P"),

    tc("FE-SEO-003","SEO","Meta","Campaign page has SEO meta",
       "1. View source of /campaigns/{id}",
       "Campaign exists",
       "<title> contains campaign title\n<meta name=description> present\nog:title, og:image, og:url present\ntwitter:card present","H","P"),

    tc("FE-SEO-004","SEO","Theme","Theme color in mobile browser",
       "1. Open on mobile Chrome",
       "Mobile device or DevTools mobile mode",
       "Browser address bar tinted purple (#7c3aed)","L","P"),

    # ── RESPONSIVE / ACCESSIBILITY ────────────────────────────────────────────
    tc("FE-RES-001","Responsive","Mobile","Home page on 375px width",
       "1. Open DevTools\n2. Set viewport to 375x812\n3. Open home page",
       "App running",
       "No horizontal scroll\nAll text readable\nButtons large enough to tap\nCampaign cards stack vertically","H","P"),

    tc("FE-RES-002","Responsive","Mobile","Entry form on mobile",
       "1. Open campaign on 375px\n2. Fill entry form",
       "Mobile viewport, active campaign",
       "Form inputs full-width\nKeyboard doesn't hide submit button\nSubmit works","H","P"),

    tc("FE-RES-003","Responsive","Tablet","Dashboard on 768px",
       "1. Set viewport to 768x1024\n2. Open dashboard",
       "Logged in user",
       "Sidebar transitions correctly\nStats grid adapts to 2-col layout","M","P"),

    tc("FE-RES-004","Responsive","Desktop","Dashboard on 1440px",
       "1. Open dashboard on 1440px viewport",
       "Logged in user",
       "Sidebar fixed on left\nContent area uses full available width\nNo oversized text","M","P"),

    # ── ERROR STATES ──────────────────────────────────────────────────────────
    tc("FE-ERR-001","Error Handling","404","Non-existent campaign shows 404",
       "1. Navigate to /campaigns/507f1f77bcf86cd799439011",
       "Campaign ID not in DB",
       "Next.js not-found page shown\nNo crash\nNavigation still works","H","N"),

    tc("FE-ERR-002","Error Handling","API Down","UI when backend is down",
       "1. Stop FastAPI server\n2. Open campaigns page",
       "Backend not running",
       "Error state shown (not blank page or JS crash)\nUser informed of connection issue","H","N"),

    tc("FE-ERR-003","Error Handling","Network","Network slow loading states",
       "1. Use DevTools Network throttling (Slow 3G)\n2. Open campaigns page",
       "Throttled network",
       "Loading spinners/skeletons shown\nNo layout shift after data loads","M","E"),
]


# ══════════════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

def build_sheet(wb, title: str, cases: list, sheet_color: str):
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = sheet_color
    ws.sheet_view.showGridLines   = False

    apply_header(ws)

    for i, row in enumerate(cases, start=2):
        write_row(ws, i, row, tc_type=row[10][0] if row[10] else "P")

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    return ws


def build_summary(wb, backend_cases, frontend_cases):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_properties.tabColor = "222222"
    ws.sheet_view.showGridLines   = False

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "GiveAwayLead — Test Case Summary"
    title_cell.font      = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill      = hex_fill("4F2D8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = "Auto-generated test case suite covering Backend API, Frontend UI, SEO, PWA, Responsive, and Error handling"
    sub.font      = Font(italic=True, size=11, color="6B6B6B")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 24

    headers = ["Category", "Module", "Total TCs", "High Priority", "Medium Priority", "Low Priority"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.fill   = hex_fill("2D1B69")
        c.font   = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[4].height = 30

    def count_group(cases, module_filter=None, priority_filter=None):
        subset = cases
        if module_filter:
            subset = [c for c in subset if c[1] == module_filter]
        if priority_filter:
            subset = [c for c in subset if c[9] == priority_filter]
        return len(subset)

    # Group by module
    all_cases = [("Backend", c) for c in backend_cases] + [("Frontend", c) for c in frontend_cases]
    modules: dict = {}
    for cat, c in all_cases:
        key = (cat, c[1])
        modules.setdefault(key, []).append(c)

    row = 5
    for (cat, mod), cases in sorted(modules.items()):
        h = sum(1 for c in cases if c[9] == "High")
        m = sum(1 for c in cases if c[9] == "Medium")
        l = sum(1 for c in cases if c[9] == "Low")
        vals = [cat, mod, len(cases), h, m, l]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
            if row % 2 == 0:
                cell.fill = hex_fill("F0EBF8")
        ws.row_dimensions[row].height = 22
        row += 1

    # Totals row
    row += 1
    totals = [
        "TOTAL", "All Modules",
        len(backend_cases) + len(frontend_cases),
        count_group(backend_cases + frontend_cases, priority_filter="High"),
        count_group(backend_cases + frontend_cases, priority_filter="Medium"),
        count_group(backend_cases + frontend_cases, priority_filter="Low"),
    ]
    for ci, v in enumerate(totals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.fill   = hex_fill("4F2D8A")
        cell.font   = Font(bold=True, color="FFFFFF")
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
    ws.row_dimensions[row].height = 26

    for ci, w in enumerate([18, 22, 14, 16, 18, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Legend
    row += 2
    ws.cell(row=row, column=1, value="Legend — Priority").font = Font(bold=True)
    for i, (label, color, fc) in enumerate([("High", COLORS["high"], "FFFFFF"), ("Medium", COLORS["medium"], "FFFFFF"), ("Low", COLORS["low"], "FFFFFF")], 1):
        c = ws.cell(row=row+1, column=i, value=label)
        c.fill = hex_fill(color); c.font = Font(bold=True, color=fc)
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()

    row += 3
    ws.cell(row=row, column=1, value="Legend — Test Type").font = Font(bold=True)
    for i, (label, color, fc) in enumerate([("Positive", "D6F5D6", "1B5E20"), ("Negative", "FFE0E0", "B71C1C"), ("Edge Case", "FFF3CC", "795548")], 1):
        c = ws.cell(row=row+1, column=i, value=label)
        c.fill = hex_fill(color); c.font = Font(bold=True, color=fc)
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_summary(wb, BACKEND_CASES, FRONTEND_CASES)
    build_sheet(wb, "Backend API Tests",  BACKEND_CASES,  "6D28D9")
    build_sheet(wb, "Frontend UI Tests",  FRONTEND_CASES, "0891B2")

    out = r"D:\mylearning\Giveway\GiveAwayLead_Test_Cases.xlsx"
    wb.save(out)
    print(f"[OK] Saved: {out}")
    print(f"   Backend TCs  : {len(BACKEND_CASES)}")
    print(f"   Frontend TCs : {len(FRONTEND_CASES)}")
    print(f"   Total        : {len(BACKEND_CASES) + len(FRONTEND_CASES)}")


if __name__ == "__main__":
    main()
